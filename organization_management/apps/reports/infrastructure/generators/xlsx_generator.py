from io import BytesIO
from typing import Tuple

import openpyxl
from openpyxl.styles import Font, Alignment


class XLSXGenerator:
    """
    Простейший XLSX‑генератор: выводит сводную таблицу по подразделениям.
    Возвращает кортеж (filename, bytes).
    """

    def generate(self, data, report) -> Tuple[str, bytes]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчет"

        ws["A1"].value = f"Отчет: {report.get_report_type_display()}"
        ws["A2"].value = f"Раздел: {data.get('division')}"
        ws["A3"].value = f"Дата: {data.get('date')}"
        ws["A1"].font = Font(bold=True)
        ws["A2"].font = Font(italic=True)

        headers = [
            "Подразделение",
            "Штатная",
            "В строю",
            "Отпуск",
            "Больничный",
            "Командировка",
            "Учёба",
            "Прикомандировано",
            "Откомандировано",
            "Прочие отсутствия",
            "Итого налич.",
            "% налич.",
        ]
        ws.append(headers)
        for cell in ws[4]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row in data.get("rows", []):
            ws.append(
                [
                    row["division_name"],
                    row["staff_unit"],
                    row["in_service"],
                    row["vacation"],
                    row["sick_leave"],
                    row["business_trip"],
                    row["training"],
                    row["seconded_in"],
                    row["seconded_out"],
                    row["other_absence"],
                    row["present_total"],
                    row["presence_pct"],
                ]
            )

        stream = BytesIO()
        wb.save(stream)
        content = stream.getvalue()
        filename = f"report_{report.id}.xlsx"
        return filename, content
