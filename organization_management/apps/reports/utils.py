"""
Утилиты для генерации отчетов
"""
import os
import logging
from datetime import datetime
from io import BytesIO
from django.db.models import Q

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from django.conf import settings

from organization_management.apps.divisions.models import Division
from organization_management.apps.staff_unit.models import StaffUnit
from organization_management.apps.employees.models import Employee
from organization_management.apps.statuses.models import EmployeeStatus

logger = logging.getLogger(__name__)


def get_active_status_filter(report_date):
    """
    Возвращает Q-фильтр для получения статусов, активных или запланированных на указанную дату.

    Статус считается активным/запланированным на дату, если:
    - start_date <= report_date
    - end_date >= report_date ИЛИ end_date is None
    - state = ACTIVE или PLANNED

    Args:
        report_date: Дата для проверки

    Returns:
        Q: Django Q-объект для фильтрации
    """
    return Q(
        start_date__lte=report_date,
        state__in=[EmployeeStatus.StatusState.ACTIVE, EmployeeStatus.StatusState.PLANNED]
    ) & (
        Q(end_date__gte=report_date) | Q(end_date__isnull=True)
    )


def get_employee_status_on_date(employee, report_date, status_type):
    """
    Получает статус сотрудника на указанную дату.
    Если нет активного статуса указанного типа на эту дату, возвращает None.

    Args:
        employee: Объект сотрудника
        report_date: Дата для проверки
        status_type: Тип статуса для поиска

    Returns:
        EmployeeStatus или None
    """
    status_filter = get_active_status_filter(report_date)
    return EmployeeStatus.objects.filter(
        employee=employee,
        status_type=status_type
    ).filter(status_filter).first()


def get_employees_by_status_on_date(division_ids, report_date, status_types):
    """
    Получает список сотрудников с указанными типами статусов на дату.

    Args:
        division_ids: ID подразделений
        report_date: Дата для проверки
        status_types: Список типов статусов (может быть одним значением или списком)

    Returns:
        QuerySet EmployeeStatus
    """
    status_filter = get_active_status_filter(report_date)

    if not isinstance(status_types, list):
        status_types = [status_types]

    return EmployeeStatus.objects.filter(
        employee__staff_unit__division_id__in=division_ids if isinstance(division_ids, list) else [division_ids],
        status_type__in=status_types
    ).filter(status_filter).select_related('employee', 'related_division')


def safe_set_cell_value(ws, row, col, value):
    """
    Безопасно устанавливает значение ячейки, даже если это объединенная ячейка.
    Записывает в первую (верхнюю левую) ячейку объединенного диапазона.
    """
    cell = ws.cell(row=row, column=col)

    # Проверяем, является ли ячейка частью объединенного диапазона
    if isinstance(cell, MergedCell):
        # Находим диапазон объединенных ячеек
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Получаем координаты первой ячейки диапазона
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                # Записываем значение в первую ячейку
                ws.cell(row=min_row, column=min_col).value = value
                logger.debug(f"Set merged cell ({row},{col}) -> ({min_row},{min_col}) = {value}")
                return
        # Если диапазон не найден, логируем предупреждение
        logger.warning(f"MergedCell at ({row},{col}) not found in merged_ranges, trying direct write")
        # Пытаемся записать напрямую
        try:
            ws.cell(row=row, column=col).value = value
        except:
            logger.error(f"Failed to write value {value} to cell ({row},{col})")
    else:
        # Обычная ячейка - просто записываем значение
        cell.value = value
        logger.debug(f"Set cell ({row},{col}) = {value}")


def generate_personnel_expense_report(department_id, report_date=None):
    """
    Генерирует отчет "Расход" по департаменту в памяти (динамически без шаблона).

    Args:
        department_id: ID департамента
        report_date: Дата для определения статусов сотрудников (datetime.date).
                     Если None, используется текущая дата.

    Returns:
        tuple: (BytesIO объект с Excel файлом, имя файла)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # Если дата не указана, используем текущую
    if report_date is None:
        from datetime import date
        report_date = date.today()

    # Создаем новый Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Расход"

    # Получаем департамент
    try:
        department = Division.objects.get(pk=department_id, division_type=Division.DivisionType.DEPARTMENT)
    except Division.DoesNotExist:
        raise ValueError(f'Департамент с ID {department_id} не найден')

    heads = Division.objects.filter(division_type=Division.DivisionType.DEPARTMENT, is_active=True).order_by('name')

    # Получаем все управления в департаменте (level=2)
    all_descendants = department.get_descendants(include_self=False)
    directorates = all_descendants.filter(
        level=2,
        is_active=True
    ).order_by('name')

    # Стили для ячеек
    header_font = Font(bold=True, size=11)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Светло-серый цвет для заголовков и итоговой строки
    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # СТРОКА 1: Название департамента
    ws.merge_cells('A1:M1')
    cell = ws['A1']
    cell.value = f"Расход личного состава за {department.name} на {report_date.strftime('%d.%m.%Y')}"
    cell.font = Font(bold=True, size=14)
    cell.alignment = center_alignment

    # СТРОКА 2: Пустая строка для отступа

    # СТРОКА 3: Заголовки таблицы
    headers = [
        'Подразделение',
        'Штат',
        'Списочный состав',
        'В строю',
        'Вакансии',
        'Отпуск',
        'Командировка',
        'Больничный',
        'На дежурстве',
        'После дежурства',
        'На учебе',
        'Прикомандирован',
        'Откомандирован'
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = gray_fill  # Светло-серый цвет

    # Устанавливаем ширину столбцов (в Excel 1 единица ширины ≈ 0.18 см)
    # A = 8 см ≈ 44 единиц
    # B, C, D, E = 2.5 см ≈ 14 единиц
    # F-M = 5 см ≈ 28 единиц
    ws.column_dimensions['A'].width = 44
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 14
    for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col].width = 28

    # Начальная строка для данных (строка 4 - Руководство)
    current_row = 4

    # Итоговые счетчики
    total_staff_units = 0
    total_employees = 0
    total_in_service = 0
    total_vacancies = 0
    total_vacation = 0
    total_sick = 0
    total_business_trip = 0
    total_on_duty = 0
    total_after_duty = 0
    total_training = 0
    total_seconded_from = 0
    total_seconded_to = 0

    # === ОБРАБОТКА РУКОВОДСТВА ДЕПАРТАМЕНТА ===
    # Получаем сотрудников, напрямую относящихся к департаменту (не к управлениям)
    head_division_ids = [department.id]

    # Количество штатных единиц руководства
    head_staff_units = StaffUnit.objects.filter(division_id=department.id).count()

    # Количество сотрудников руководства
    head_employees = StaffUnit.objects.filter(
        division_id=department.id,
        employee__isnull=False
    ).count()

    # Получаем фильтр для активных статусов на дату
    status_filter = get_active_status_filter(report_date)

    # Получаем сотрудников с любыми статусами кроме "в строю" на эту дату
    employees_with_other_statuses = EmployeeStatus.objects.filter(
        employee__staff_unit__division_id=department.id
    ).filter(status_filter).exclude(
        status_type=EmployeeStatus.StatusType.IN_SERVICE
    ).values_list('employee_id', flat=True).distinct()

    # "В строю" = Все сотрудники - Сотрудники с другими статусами
    head_in_service = head_employees - len(employees_with_other_statuses)

    # Вакансии руководства
    head_vacancies = StaffUnit.objects.filter(
        division_id=department.id,
        employee__isnull=True
    ).count()

    # Отпуск руководства
    head_vacation_statuses = get_employees_by_status_on_date(
        department.id,
        report_date,
        [EmployeeStatus.StatusType.VACATION, EmployeeStatus.StatusType.LEAVE_BY_REPORT]
    )
    head_vacation_count = head_vacation_statuses.count()
    head_vacation_list = []
    for status in head_vacation_statuses:
        emp = status.employee
        period = f"{status.start_date.strftime('%d.%m.%Y') if status.start_date else ''} - {status.end_date.strftime('%d.%m.%Y') if status.end_date else ''}"
        head_vacation_list.append(f"{emp.last_name} {emp.first_name} ({period})")

    # Командировка руководства
    head_trip_statuses = get_employees_by_status_on_date(
        department.id,
        report_date,
        EmployeeStatus.StatusType.BUSINESS_TRIP
    )
    head_trip_count = head_trip_statuses.count()
    head_trip_list = []
    for status in head_trip_statuses:
        emp = status.employee
        period = f"{status.start_date.strftime('%d.%m.%Y') if status.start_date else ''} - {status.end_date.strftime('%d.%m.%Y') if status.end_date else ''}"
        head_trip_list.append(f"{emp.last_name} {emp.first_name} ({period})")

    # Больничный руководства
    head_sick_statuses = get_employees_by_status_on_date(
        department.id,
        report_date,
        EmployeeStatus.StatusType.SICK_LEAVE
    )
    head_sick_count = head_sick_statuses.count()
    head_sick_list = []
    for status in head_sick_statuses:
        emp = status.employee
        period = f"{status.start_date.strftime('%d.%m.%Y') if status.start_date else ''} - {status.end_date.strftime('%d.%m.%Y') if status.end_date else ''}"
        head_sick_list.append(f"{emp.last_name} {emp.first_name} ({period})")

    # На дежурстве руководства
    head_on_duty_statuses = get_employees_by_status_on_date(
        department.id,
        report_date,
        EmployeeStatus.StatusType.ON_DUTY
    )
    head_on_duty_count = head_on_duty_statuses.count()
    head_on_duty_list = [f"{s.employee.last_name} {s.employee.first_name}" for s in head_on_duty_statuses]

    # После дежурства руководства
    head_after_duty_statuses = get_employees_by_status_on_date(
        department.id,
        report_date,
        EmployeeStatus.StatusType.AFTER_DUTY
    )
    head_after_duty_count = head_after_duty_statuses.count()
    head_after_duty_list = [f"{s.employee.last_name} {s.employee.first_name}" for s in head_after_duty_statuses]

    # На учебе руководства
    head_training_statuses = get_employees_by_status_on_date(
        department.id,
        report_date,
        [EmployeeStatus.StatusType.TRAINING, EmployeeStatus.StatusType.COMPETITION]
    )
    head_training_count = head_training_statuses.count()
    head_training_list = []
    for status in head_training_statuses:
        emp = status.employee
        period = f"{status.start_date.strftime('%d.%m.%Y') if status.start_date else ''} - {status.end_date.strftime('%d.%m.%Y') if status.end_date else ''}"
        head_training_list.append(f"{emp.last_name} {emp.first_name} ({period})")

    def _dep_name(div):
        if not div:
            return "?"
        dep = div
        while dep and dep.level > 1:
            dep = dep.parent
        return dep.name if dep else div.name

    # Прикомандирован руководства
    head_seconded_from_statuses = get_employees_by_status_on_date(
        department.id,
        report_date,
        EmployeeStatus.StatusType.SECONDED_FROM
    )
    head_seconded_from_count = head_seconded_from_statuses.count()
    head_seconded_from_list = []
    for status in head_seconded_from_statuses:
        emp = status.employee
        period = f"{status.start_date.strftime('%d.%m.%Y') if status.start_date else ''} - {status.end_date.strftime('%d.%m.%Y') if status.end_date else ''}"
        from_div = _dep_name(status.related_division)
        head_seconded_from_list.append(f"{emp.last_name} {emp.first_name} ({period}, из {from_div})")

    # Откомандирован руководства
    head_seconded_to_statuses = get_employees_by_status_on_date(
        department.id,
        report_date,
        EmployeeStatus.StatusType.SECONDED_TO
    )
    head_seconded_to_count = head_seconded_to_statuses.count()
    head_seconded_to_list = []
    for status in head_seconded_to_statuses:
        emp = status.employee
        period = f"{status.start_date.strftime('%d.%м.%Y') if status.start_date else ''} - {status.end_date.strftime('%d.%м.%Y') if status.end_date else ''}"
        to_div = _dep_name(status.related_division)
        head_seconded_to_list.append(f"{emp.last_name} {emp.first_name} ({period}, в {to_div})")

    # СТРОКА 4: Руководство - Название + Числа
    head_start_row = current_row

    row_data = [
        "Басшылық",  # "Руководство"
        head_staff_units,
        head_employees,
        head_in_service,
        head_vacancies,
        head_vacation_count,
        head_trip_count,
        head_sick_count,
        head_on_duty_count,
        head_after_duty_count,
        head_training_count,
        head_seconded_from_count,
        head_seconded_to_count
    ]

    for col_num, value in enumerate(row_data, start=1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal='left' if col_num == 1 else 'center', vertical='center')

    # СТРОКА 5: Руководство - Подробности ФИО
    current_row += 1

    detail_data = {
        2: "",  # Пустые колонки B-E
        3: "",
        4: "",
        5: "",
        6: "; ".join(head_vacation_list) if head_vacation_list else "",
        7: "; ".join(head_trip_list) if head_trip_list else "",
        8: "; ".join(head_sick_list) if head_sick_list else "",
        9: "; ".join(head_on_duty_list) if head_on_duty_list else "",
        10: "; ".join(head_after_duty_list) if head_after_duty_list else "",
        11: "; ".join(head_training_list) if head_training_list else "",
        12: "; ".join(head_seconded_from_list) if head_seconded_from_list else "",
        13: "; ".join(head_seconded_to_list) if head_seconded_to_list else ""
    }

    for col_num, value in detail_data.items():
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.font = Font(size=9)

    # Устанавливаем высоту строки с деталями (4 см ≈ 113 пунктов)
    ws.row_dimensions[current_row].height = 113

    # Объединяем ячейки A, B, C, D, E по вертикали (строки 4 и 5)
    for col in range(1, 6):  # Колонки A-E (1-5)
        ws.merge_cells(start_row=head_start_row, start_column=col, end_row=current_row, end_column=col)
        cell = ws.cell(row=head_start_row, column=col)
        cell.alignment = Alignment(horizontal='left' if col == 1 else 'center', vertical='center')
        if col == 1:
            cell.font = Font(bold=True)

    # Добавляем руководство к итоговым счетчикам
    total_staff_units += head_staff_units
    total_employees += head_employees
    total_in_service += head_in_service
    total_vacancies += head_vacancies
    total_vacation += head_vacation_count
    total_business_trip += head_trip_count
    total_sick += head_sick_count
    total_on_duty += head_on_duty_count
    total_after_duty += head_after_duty_count
    total_training += head_training_count
    total_seconded_from += head_seconded_from_count
    total_seconded_to += head_seconded_to_count

    # Обрабатываем каждое управление
    for directorate in directorates:
        # Получаем всех сотрудников управления (включая подчиненные отделы)
        directorate_descendants = directorate.get_descendants(include_self=True)
        directorate_division_ids = list(directorate_descendants.values_list('id', flat=True))

        # Количество штатных единиц
        staff_units_count = StaffUnit.objects.filter(
            division_id__in=directorate_division_ids
        ).count()
        total_staff_units += staff_units_count

        # Количество сотрудников
        employees_count = StaffUnit.objects.filter(
            division_id__in=directorate_division_ids,
            employee__isnull=False
        ).count()
        total_employees += employees_count

        # Получаем сотрудников с любыми статусами кроме "в строю" на эту дату
        dir_employees_with_other_statuses = EmployeeStatus.objects.filter(
            employee__staff_unit__division_id__in=directorate_division_ids
        ).filter(status_filter).exclude(
            status_type=EmployeeStatus.StatusType.IN_SERVICE
        ).values_list('employee_id', flat=True).distinct()

        # "В строю" = Все сотрудники - Сотрудники с другими статусами
        in_service_count = employees_count - len(dir_employees_with_other_statuses)
        total_in_service += in_service_count

        # Количество вакансий
        vacancies_count = StaffUnit.objects.filter(
            division_id__in=directorate_division_ids,
            employee__isnull=True
        ).count()
        total_vacancies += vacancies_count

        # Отпуск
        vacation_statuses = get_employees_by_status_on_date(
            directorate_division_ids,
            report_date,
            [EmployeeStatus.StatusType.VACATION, EmployeeStatus.StatusType.LEAVE_BY_REPORT]
        )
        vacation_count = vacation_statuses.count()
        vacation_list = []
        for status in vacation_statuses:
            emp = status.employee
            period = f"{status.start_date.strftime('%d.%m.%Y') if status.start_date else ''} - {status.end_date.strftime('%d.%m.%Y') if status.end_date else ''}"
            vacation_list.append(f"{emp.last_name} {emp.first_name} ({period})")
        total_vacation += vacation_count

        # Командировка
        trip_statuses = get_employees_by_status_on_date(
            directorate_division_ids,
            report_date,
            EmployeeStatus.StatusType.BUSINESS_TRIP
        )
        trip_count = trip_statuses.count()
        trip_list = []
        for status in trip_statuses:
            emp = status.employee
            period = f"{status.start_date.strftime('%d.%m.%Y') if status.start_date else ''} - {status.end_date.strftime('%d.%m.%Y') if status.end_date else ''}"
            trip_list.append(f"{emp.last_name} {emp.first_name} ({period})")
        total_business_trip += trip_count

        # Больничный
        sick_statuses = get_employees_by_status_on_date(
            directorate_division_ids,
            report_date,
            EmployeeStatus.StatusType.SICK_LEAVE
        )
        sick_count = sick_statuses.count()
        sick_list = []
        for status in sick_statuses:
            emp = status.employee
            period = f"{status.start_date.strftime('%d.%m.%Y') if status.start_date else ''} - {status.end_date.strftime('%d.%m.%Y') if status.end_date else ''}"
            sick_list.append(f"{emp.last_name} {emp.first_name} ({period})")
        total_sick += sick_count

        # На дежурстве
        on_duty_statuses = get_employees_by_status_on_date(
            directorate_division_ids,
            report_date,
            EmployeeStatus.StatusType.ON_DUTY
        )
        on_duty_count = on_duty_statuses.count()
        on_duty_list = [f"{s.employee.last_name} {s.employee.first_name}" for s in on_duty_statuses]
        total_on_duty += on_duty_count

        # После дежурства
        after_duty_statuses = get_employees_by_status_on_date(
            directorate_division_ids,
            report_date,
            EmployeeStatus.StatusType.AFTER_DUTY
        )
        after_duty_count = after_duty_statuses.count()
        after_duty_list = [f"{s.employee.last_name} {s.employee.first_name}" for s in after_duty_statuses]
        total_after_duty += after_duty_count

        # На учебе/соревнованиях
        training_statuses = get_employees_by_status_on_date(
            directorate_division_ids,
            report_date,
            [EmployeeStatus.StatusType.TRAINING, EmployeeStatus.StatusType.COMPETITION]
        )
        training_count = training_statuses.count()
        training_list = []
        for status in training_statuses:
            emp = status.employee
            period = f"{status.start_date.strftime('%d.%m.%Y') if status.start_date else ''} - {status.end_date.strftime('%d.%m.%Y') if status.end_date else ''}"
            training_list.append(f"{emp.last_name} {emp.first_name} ({period})")
        total_training += training_count

        # Определяем текущий департамент для этого управления
        current_department = (
            directorate.parent if directorate.parent and directorate.parent.level == 1
            else directorate.get_ancestors(include_self=True).filter(level=1).first()
        )

        def _resolve_division_name(div):
            """Если из нашего департамента — возвращаем управление, иначе имя департамента."""
            if not div:
                return "?"
            dep = div
            while dep and dep.level > 1:
                dep = dep.parent

            # Наш департамент: показываем управление
            if current_department and dep and dep.id == current_department.id:
                node = div
                while node and node.level > 2:
                    node = node.parent
                return node.name if node else div.name  # управление (level=2) или исходное имя

            # Чужой департамент: название департамента
            return dep.name if dep else div.name

        # Активный фильтр по датам/состоянию
        status_filter = get_active_status_filter(report_date)

        # Прикомандированы в наше управление (related_division ∈ наше управление; сотрудник из другого управления/департамента)
        seconded_in_statuses = (
            EmployeeStatus.objects
            .filter(
                status_type=EmployeeStatus.StatusType.SECONDED_TO,        # статус хранит целевое подразделение в related_division
                related_division_id__in=directorate_division_ids,          # целевое подразделение — наше управление/подразделения
            )
            .exclude(employee__staff_unit__division_id__in=directorate_division_ids)  # сотрудник не из нашего управления
            .filter(status_filter)
            .select_related('employee', 'related_division')
        )

        seconded_from_count = seconded_in_statuses.count()
        seconded_from_list = []
        for status in seconded_in_statuses:
            emp = status.employee
            period = f"{status.start_date.strftime('%d.%m.%Y') if status.start_date else ''} - {status.end_date.strftime('%d.%m.%Y') if status.end_date else ''}"
            source_div = emp.staff_unit.division if getattr(emp, "staff_unit", None) and emp.staff_unit.division else None
            from_div = _resolve_division_name(source_div)
            seconded_from_list.append(f"{emp.last_name} {emp.first_name} ({period}, из {from_div})")
        total_seconded_from += seconded_from_count

        # Откомандированы из нашего управления (staff_unit ∈ наше управление; related_division — любое другое подразделение)
        seconded_out_statuses = (
            EmployeeStatus.objects
            .filter(
                status_type=EmployeeStatus.StatusType.SECONDED_TO,         # тот же тип, но смотрим на исходящее
                employee__staff_unit__division_id__in=directorate_division_ids,
            )
            .exclude(related_division_id__in=directorate_division_ids)     # целевое подразделение не наше
            .filter(status_filter)
            .select_related('employee', 'related_division')
        )

        seconded_to_count = seconded_out_statuses.count()
        seconded_to_list = []
        for status in seconded_out_statuses:
            emp = status.employee
            period = f"{status.start_date.strftime('%d.%m.%Y') if status.start_date else ''} - {status.end_date.strftime('%d.%m.%Y') if status.end_date else ''}"
            target_div = status.related_division if status.related_division else None
            to_div = _resolve_division_name(target_div)
            seconded_to_list.append(f"{emp.last_name} {emp.first_name} ({period}, в {to_div})")
        total_seconded_to += seconded_to_count

        # СТРОКА: Управление - Название + Числа
        current_row += 1
        dir_start_row = current_row

        row_data = [
            directorate.name,
            staff_units_count,
            employees_count,
            in_service_count,
            vacancies_count,
            vacation_count,
            trip_count,
            sick_count,
            on_duty_count,
            after_duty_count,
            training_count,
            seconded_from_count,
            seconded_to_count
        ]

        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left' if col_num == 1 else 'center', vertical='center')

        # СТРОКА: Подробности ФИО
        current_row += 1

        detail_data = {
            2: "",  # Пустые колонки B-E
            3: "",
            4: "",
            5: "",
            6: "; ".join(vacation_list) if vacation_list else "",
            7: "; ".join(trip_list) if trip_list else "",
            8: "; ".join(sick_list) if sick_list else "",
            9: "; ".join(on_duty_list) if on_duty_list else "",
            10: "; ".join(after_duty_list) if after_duty_list else "",
            11: "; ".join(training_list) if training_list else "",
            12: "; ".join(seconded_from_list) if seconded_from_list else "",
            13: "; ".join(seconded_to_list) if seconded_to_list else ""
        }

        for col_num, value in detail_data.items():
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.font = Font(size=9)

        # Устанавливаем высоту строки с деталями (4 см ≈ 113 пунктов)
        ws.row_dimensions[current_row].height = 113

        # Объединяем ячейки A, B, C, D, E по вертикали (название управления на 2 строки)
        for col in range(1, 6):  # Колонки A-E (1-5)
            ws.merge_cells(start_row=dir_start_row, start_column=col, end_row=current_row, end_column=col)
            cell = ws.cell(row=dir_start_row, column=col)
            cell.alignment = Alignment(horizontal='left' if col == 1 else 'center', vertical='center')
            if col == 1:
                cell.font = Font(bold=True)

    # ИТОГОВАЯ СТРОКА
    current_row += 1

    total_data = [
        "ИТОГО",
        total_staff_units,
        total_employees,
        total_in_service,
        total_vacancies,
        total_vacation,
        total_business_trip,
        total_sick,
        total_on_duty,
        total_after_duty,
        total_training,
        total_seconded_from,
        total_seconded_to
    ]

    for col_num, value in enumerate(total_data, start=1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal='left' if col_num == 1 else 'center', vertical='center')
        cell.font = Font(bold=True)
        cell.fill = gray_fill  # Светло-серый цвет

    # Сохраняем в память
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Формируем имя файла с датой отчета
    date_str = report_date.strftime('%Y-%m-%d')
    filename = f"расход_{department.name}_на_{date_str}.xlsx"

    return output, filename


def generate_organization_report(report_date=None):
    """
    Генерирует отчет "Организация" по всем департаментам в памяти (динамически).

    Args:
        report_date: Дата для определения статусов сотрудников (datetime.date).
                     Если None, используется текущая дата.

    Returns:
        tuple: (BytesIO объект с Excel файлом, имя файла)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # Если дата не указана, используем текущую
    if report_date is None:
        from datetime import date
        report_date = date.today()

    # Создаем новый Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Организация"

    # Получаем все активные департаменты (level=1)
    departments = Division.objects.filter(
        level=1,
        division_type=Division.DivisionType.DEPARTMENT,
        is_active=True
    ).order_by('name')

    if not departments.exists():
        raise ValueError('Департаменты не найдены')

    # Стили для ячеек
    header_font = Font(bold=True, size=11)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # СТРОКА 1: Заголовок
    ws.merge_cells('E1:P1')
    cell = ws['E1']
    cell.value = "Организация бөлімшелері бойынша саптық жазба"
    cell.font = Font(bold=True, size=14)
    cell.alignment = center_alignment

    # СТРОКА 2: Дата
    ws.merge_cells('E2:P2')
    cell = ws['E2']
    # Форматируем дату на казахском
    from datetime import datetime
    date_kz = datetime.strptime(str(report_date), '%Y-%m-%d')
    months_kz = {
        1: 'қаңтар', 2: 'ақпан', 3: 'наурыз', 4: 'сәуір', 5: 'мамыр', 6: 'маусым',
        7: 'шілде', 8: 'тамыз', 9: 'қыркүйек', 10: 'қазан', 11: 'қараша', 12: 'желтоқсан'
    }
    cell.value = f"{date_kz.year} жылғы {date_kz.day} {months_kz[date_kz.month]}ға"
    cell.font = Font(bold=True, size=12)
    cell.alignment = center_alignment

    # СТРОКА 3: Пустая для отступа

    # СТРОКА 4: Заголовки таблицы
    headers = [
        ('A', '№ р/с'),
        ('B', 'Бөлімшелер'),
        ('D', 'Тізімге сәйкес'),
        ('E', 'Сапта'),
        ('F', 'Саптан тыс'),
        ('G', 'Аты, жөні'),
        ('J', 'Демалыста'),
        ('K', 'Бала күтімі б/ша'),
        ('L', 'Іссапарда'),
        ('M', 'Изоляция'),
        ('N', 'Ауруы б/ша'),
        ('O', 'Жарыс жиын/ Оқуда'),
        ('P', 'Баянат б/ша'),
        ('Q', 'Жасақта'),
        ('R', 'Тағылымдама')
    ]

    # Устанавливаем заголовки
    for col_letter, header_text in headers:
        col_num = ord(col_letter) - ord('A') + 1
        cell = ws.cell(row=4, column=col_num)
        cell.value = header_text
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = gray_fill

    # Объединяем ячейки для заголовков
    ws.merge_cells('B4:C4')  # Бөлімшелер
    ws.merge_cells('G4:I4')  # Аты, жөні

    # Устанавливаем ширину столбцов
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 5
    for col in ['D', 'E', 'F']:
        ws.column_dimensions[col].width = 12
    for col in ['G', 'H', 'I']:
        ws.column_dimensions[col].width = 15
    for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        ws.column_dimensions[col].width = 12

    # Начальная строка для данных
    current_row = 5

    # Итоговые счетчики
    total_staff = 0
    total_in_service = 0
    total_out_of_service = 0
    total_vacation = 0
    total_child_care = 0
    total_business_trip = 0
    total_isolation = 0
    total_sick = 0
    total_training_competition = 0
    total_leave_by_report = 0
    total_on_duty = 0
    total_internship = 0

    # Список ФИО руководства с нестандартными статусами
    all_leadership_fio = []

    # Фильтр для активных статусов
    status_filter = get_active_status_filter(report_date)

    # === ОБРАБОТКА РУКОВОДСТВА ОРГАНИЗАЦИИ (корневое подразделение) ===
    # Получаем корневое подразделение (СГО РК, level=0)
    root_division = Division.objects.filter(level=0, is_active=True).first()

    if root_division:
        row_num = 1

        # Количество штатных единиц руководства
        org_staff_count = StaffUnit.objects.filter(division_id=root_division.id).count()
        total_staff += org_staff_count

        # Количество сотрудников руководства
        org_employees_count = StaffUnit.objects.filter(
            division_id=root_division.id,
            employee__isnull=False
        ).count()

        # Получаем сотрудников с любыми статусами кроме "в строю"
        org_employees_with_other_statuses = EmployeeStatus.objects.filter(
            employee__staff_unit__division_id=root_division.id
        ).filter(status_filter).exclude(
            status_type=EmployeeStatus.StatusType.IN_SERVICE
        ).values_list('employee_id', flat=True).distinct()

        # "В строю" = Все сотрудники - Сотрудники с другими статусами
        org_in_service_count = org_employees_count - len(org_employees_with_other_statuses)
        total_in_service += org_in_service_count

        # "Саптан тыс" (вне строя)
        org_out_of_service_count = len(org_employees_with_other_statuses)
        total_out_of_service += org_out_of_service_count

        # Получаем руководство с нестандартными статусами
        org_leadership_statuses = EmployeeStatus.objects.filter(
            employee__staff_unit__division_id=root_division.id
        ).filter(status_filter).exclude(
            status_type=EmployeeStatus.StatusType.IN_SERVICE
        ).select_related('employee')

        org_leadership_fio_list = []
        for status in org_leadership_statuses:
            emp = status.employee
            status_label = status.get_status_type_display()
            org_leadership_fio_list.append(f"{emp.last_name} {emp.first_name} ({status_label})")

        org_leadership_fio = "; ".join(org_leadership_fio_list) if org_leadership_fio_list else ""
        if org_leadership_fio_list:
            all_leadership_fio.extend(org_leadership_fio_list)

        # Считаем статусы по типам для руководства организации
        org_vacation_count = get_employees_by_status_on_date(
            root_division.id,
            report_date,
            EmployeeStatus.StatusType.VACATION
        ).count()
        total_vacation += org_vacation_count

        org_child_care_count = get_employees_by_status_on_date(
            root_division.id,
            report_date,
            EmployeeStatus.StatusType.CHILD_CARE
        ).count()
        total_child_care += org_child_care_count

        org_business_trip_count = get_employees_by_status_on_date(
            root_division.id,
            report_date,
            EmployeeStatus.StatusType.BUSINESS_TRIP
        ).count()
        total_business_trip += org_business_trip_count

        org_isolation_count = get_employees_by_status_on_date(
            root_division.id,
            report_date,
            EmployeeStatus.StatusType.ISOLATION
        ).count()
        total_isolation += org_isolation_count

        org_sick_count = get_employees_by_status_on_date(
            root_division.id,
            report_date,
            EmployeeStatus.StatusType.SICK_LEAVE
        ).count()
        total_sick += org_sick_count

        org_training_competition_count = get_employees_by_status_on_date(
            root_division.id,
            report_date,
            [EmployeeStatus.StatusType.TRAINING, EmployeeStatus.StatusType.COMPETITION]
        ).count()
        total_training_competition += org_training_competition_count

        org_leave_by_report_count = get_employees_by_status_on_date(
            root_division.id,
            report_date,
            EmployeeStatus.StatusType.LEAVE_BY_REPORT
        ).count()
        total_leave_by_report += org_leave_by_report_count

        org_on_duty_count = get_employees_by_status_on_date(
            root_division.id,
            report_date,
            EmployeeStatus.StatusType.ON_DUTY
        ).count()
        total_on_duty += org_on_duty_count

        org_internship_count = get_employees_by_status_on_date(
            root_division.id,
            report_date,
            EmployeeStatus.StatusType.INTERNSHIP
        ).count()
        total_internship += org_internship_count

        # Записываем данные руководства организации
        row_data = {
            1: row_num,  # № р/с
            2: "Организация басшылығы",  # Руководство организации
            4: org_staff_count,  # Тізімге сәйкес
            5: org_in_service_count,  # Сапта
            6: org_out_of_service_count,  # Саптан тыс
            7: org_leadership_fio,  # Аты, жөні
            10: org_vacation_count,  # Демалыста
            11: org_child_care_count,  # Бала күтімі б/ша
            12: org_business_trip_count,  # Іссапарда
            13: org_isolation_count,  # Изоляция
            14: org_sick_count,  # Ауруы б/ша
            15: org_training_competition_count,  # Жарыс жиын/ Оқуда
            16: org_leave_by_report_count,  # Баянат б/ша
            17: org_on_duty_count,  # Жасақта
            18: org_internship_count  # Тағылымдама
        }

        for col_num, value in row_data.items():
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [1, 4, 5, 6, 10, 11, 12, 13, 14, 15, 16, 17, 18]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

        # Объединяем ячейки для руководства организации
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws.merge_cells(f'G{current_row}:I{current_row}')

        current_row += 1
        row_num += 1
    else:
        # Если нет корневого подразделения, начинаем с 1
        row_num = 1

    # Обрабатываем каждый департамент
    for department in departments:
        # Получаем всех сотрудников департамента (включая подразделения)
        dept_descendants = department.get_descendants(include_self=True)
        dept_division_ids = list(dept_descendants.values_list('id', flat=True))

        # Количество штатных единиц
        staff_count = StaffUnit.objects.filter(
            division_id__in=dept_division_ids
        ).count()
        total_staff += staff_count

        # Количество сотрудников
        employees_count = StaffUnit.objects.filter(
            division_id__in=dept_division_ids,
            employee__isnull=False
        ).count()

        # Получаем сотрудников с любыми статусами кроме "в строю"
        employees_with_other_statuses = EmployeeStatus.objects.filter(
            employee__staff_unit__division_id__in=dept_division_ids
        ).filter(status_filter).exclude(
            status_type=EmployeeStatus.StatusType.IN_SERVICE
        ).values_list('employee_id', flat=True).distinct()

        # "В строю" = Все сотрудники - Сотрудники с другими статусами
        in_service_count = employees_count - len(employees_with_other_statuses)
        total_in_service += in_service_count

        # "Саптан тыс" (вне строя) = количество с нестандартными статусами
        out_of_service_count = len(employees_with_other_statuses)
        total_out_of_service += out_of_service_count

        # Получаем руководство департамента (сотрудники напрямую в департаменте, не в управлениях)
        # с нестандартными статусами
        leadership_statuses = EmployeeStatus.objects.filter(
            employee__staff_unit__division_id=department.id
        ).filter(status_filter).exclude(
            status_type=EmployeeStatus.StatusType.IN_SERVICE
        ).select_related('employee')

        leadership_fio_list = []
        for status in leadership_statuses:
            emp = status.employee
            status_label = status.get_status_type_display()
            leadership_fio_list.append(f"{emp.last_name} {emp.first_name} ({status_label})")

        leadership_fio = "; ".join(leadership_fio_list) if leadership_fio_list else ""
        if leadership_fio_list:
            all_leadership_fio.extend(leadership_fio_list)

        # Считаем статусы по типам
        # Демалыста (отпуск)
        vacation_count = get_employees_by_status_on_date(
            dept_division_ids,
            report_date,
            EmployeeStatus.StatusType.VACATION
        ).count()
        total_vacation += vacation_count

        # Бала күтімі б/ша (уход за ребенком)
        child_care_count = get_employees_by_status_on_date(
            dept_division_ids,
            report_date,
            EmployeeStatus.StatusType.CHILD_CARE
        ).count()
        total_child_care += child_care_count

        # Іссапарда (командировка)
        business_trip_count = get_employees_by_status_on_date(
            dept_division_ids,
            report_date,
            EmployeeStatus.StatusType.BUSINESS_TRIP
        ).count()
        total_business_trip += business_trip_count

        # Изоляция
        isolation_count = get_employees_by_status_on_date(
            dept_division_ids,
            report_date,
            EmployeeStatus.StatusType.ISOLATION
        ).count()
        total_isolation += isolation_count

        # Ауруы б/ша (больничный)
        sick_count = get_employees_by_status_on_date(
            dept_division_ids,
            report_date,
            EmployeeStatus.StatusType.SICK_LEAVE
        ).count()
        total_sick += sick_count

        # Жарыс жиын/ Оқуда (соревнования/учеба)
        training_competition_count = get_employees_by_status_on_date(
            dept_division_ids,
            report_date,
            [EmployeeStatus.StatusType.TRAINING, EmployeeStatus.StatusType.COMPETITION]
        ).count()
        total_training_competition += training_competition_count

        # Баянат б/ша (отпуск по рапорту)
        leave_by_report_count = get_employees_by_status_on_date(
            dept_division_ids,
            report_date,
            EmployeeStatus.StatusType.LEAVE_BY_REPORT
        ).count()
        total_leave_by_report += leave_by_report_count

        # Жасақта (дежурство)
        on_duty_count = get_employees_by_status_on_date(
            dept_division_ids,
            report_date,
            EmployeeStatus.StatusType.ON_DUTY
        ).count()
        total_on_duty += on_duty_count

        # Тағылымдама (стажировка)
        internship_count = get_employees_by_status_on_date(
            dept_division_ids,
            report_date,
            EmployeeStatus.StatusType.INTERNSHIP
        ).count()
        total_internship += internship_count

        # Записываем данные департамента
        row_data = {
            1: row_num,  # № р/с
            2: department.name,  # Бөлімшелер (B-C merged)
            4: staff_count,  # Тізімге сәйкес
            5: in_service_count,  # Сапта
            6: out_of_service_count,  # Саптан тыс
            7: leadership_fio,  # Аты, жөні (G-I merged)
            10: vacation_count,  # Демалыста
            11: child_care_count,  # Бала күтімі б/ша
            12: business_trip_count,  # Іссапарда
            13: isolation_count,  # Изоляция
            14: sick_count,  # Ауруы б/ша
            15: training_competition_count,  # Жарыс жиын/ Оқуда
            16: leave_by_report_count,  # Баянат б/ша
            17: on_duty_count,  # Жасақта
            18: internship_count  # Тағылымдама
        }

        for col_num, value in row_data.items():
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [1, 4, 5, 6, 10, 11, 12, 13, 14, 15, 16, 17, 18]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

        # Объединяем ячейки для департамента и ФИО
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws.merge_cells(f'G{current_row}:I{current_row}')

        current_row += 1
        row_num += 1

    # ИТОГОВАЯ СТРОКА (БАРЛЫҒЫ)
    ws.cell(row=current_row, column=1).value = ""
    ws.cell(row=current_row, column=2).value = "БАРЛЫҒЫ"
    ws.cell(row=current_row, column=4).value = total_staff
    ws.cell(row=current_row, column=5).value = total_in_service
    ws.cell(row=current_row, column=6).value = total_out_of_service
    ws.cell(row=current_row, column=7).value = "; ".join(all_leadership_fio) if all_leadership_fio else ""
    ws.cell(row=current_row, column=10).value = total_vacation
    ws.cell(row=current_row, column=11).value = total_child_care
    ws.cell(row=current_row, column=12).value = total_business_trip
    ws.cell(row=current_row, column=13).value = total_isolation
    ws.cell(row=current_row, column=14).value = total_sick
    ws.cell(row=current_row, column=15).value = total_training_competition
    ws.cell(row=current_row, column=16).value = total_leave_by_report
    ws.cell(row=current_row, column=17).value = total_on_duty
    ws.cell(row=current_row, column=18).value = total_internship

    # Применяем стили к итоговой строке
    for col_num in [1, 2, 4, 5, 6, 7, 10, 11, 12, 13, 14, 15, 16, 17, 18]:
        cell = ws.cell(row=current_row, column=col_num)
        cell.border = border
        cell.font = Font(bold=True)
        cell.fill = gray_fill
        if col_num in [1, 4, 5, 6, 10, 11, 12, 13, 14, 15, 16, 17, 18]:
            cell.alignment = center_alignment
        else:
            cell.alignment = left_alignment

    # Объединяем ячейки для итоговой строки
    ws.merge_cells(f'B{current_row}:C{current_row}')
    ws.merge_cells(f'G{current_row}:I{current_row}')

    # Сохраняем в память
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Формируем имя файла с датой отчета
    date_str = report_date.strftime('%Y-%m-%d')
    filename = f"организация_на_{date_str}.xlsx"

    return output, filename
